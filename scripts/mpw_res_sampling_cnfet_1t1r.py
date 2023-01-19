"""Script to perform a read voltage sweep on a chip"""
import argparse
import numpy as np
from datetime import datetime
from json import JSONEncoder
import json
import os
from itertools import product
from openpyxl import Workbook
from digitalpattern.nirram import NIRRAM

# Get arguments
parser = argparse.ArgumentParser(description="RESET a chip.")
parser.add_argument("chip", help="chip name for logging")
parser.add_argument("device", help="device name for logging")

args = parser.parse_args()

# create data folder
timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
data_foldername = f"{timestamp}_{str(args.chip)}_{str(args.device)}_res_sampling"
path_data_folder = os.path.join("data", data_foldername)
os.makedirs(path_data_folder, exist_ok=True)

# Initialize NI system
# For CNFET: make sure polarity is PMOS
nisys = NIRRAM(args.chip, args.device, settings="settings/MPW_GAX1_CNFET_1T1R_LCH160.toml", polarity="PMOS") # FOR CNFET RRAM


nisys.read(record=True)
# input("Dynamic Form")
# nisys.dynamic_form()
# nisys.dynamic_reset()

res_samples = nisys.sample_resistance_at_bias(
    mode = "SET_SAMPLE",
    initialize_cell_fn = nisys.dynamic_reset, # use bound method, either nisys.dynamic_reset or nisys.dynamic_set
    pulse_fn = nisys.set_pulse, # use bound method, either nisys.reset_pulse or nisys.set_pulse
    samples = 20,
    pw = 1,
    debug = True,
)

# keep device in reset state afterwards
nisys.dynamic_reset()

nisys.close()

# save raw res samples at bias data into json
# NOTE: must re-process resistance tuple keys since these cannot be serialized
res_samples_formatted_for_json = res_samples.copy()
res_samples_list = []
for vbl, vsl, vwl in res_samples_formatted_for_json["res"].keys():
    res_statistics = res_samples["res"][vbl, vsl, vwl]
    res_samples_list.append({
        "vbl": vbl,
        "vsl": vsl,
        "vwl": vwl,
        "res_mean": res_statistics["mean"],
        "res_median": res_statistics["median"],
        "res_std": res_statistics["std"],
        "res_values": res_statistics["values"],
    })
res_samples_formatted_for_json["res"] = res_samples_list

# https://stackoverflow.com/questions/44302423/efficient-way-to-write-a-python-dictionary-with-numpy-nd-array-values-into-a-jso
class NumpyArrayEncoder(JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return JSONEncoder.default(self, obj)

with open(os.path.join(path_data_folder, "res_samples.json"), "w+") as f:
    json.dump(res_samples_formatted_for_json, f, cls=NumpyArrayEncoder)

# get sweeps used
vbl_sweep = res_samples["vbl_sweep"]
vsl_sweep = res_samples["vsl_sweep"]
vwl_sweep = res_samples["vwl_sweep"]

# excel spreadsheet with resistance statistics calculated
wb = Workbook()
ws = wb.create_sheet("rram res statistics")
del wb["Sheet"]

# sweep sampling mode name
ws.cell(row=1, column=1).value = res_samples["mode"]

# samples
ws.cell(row=2, column=1).value = "samples"
ws.cell(row=2, column=2).value = res_samples["samples"]

# settling time
ws.cell(row=3, column=1).value = "settling_time"
ws.cell(row=3, column=2).value = res_samples["settling_time"] if res_samples["settling_time"] is not None else 0

# target hrs and lrs (for tracking)
ws.cell(row=4, column=1).value = "target_res_hrs"
ws.cell(row=4, column=2).value = res_samples["target_res_hrs"]
ws.cell(row=5, column=1).value = "target_res_lrs"
ws.cell(row=5, column=2).value = res_samples["target_res_lrs"]

# pulse width
ws.cell(row=6, column=1).value = "pw"
ws.cell(row=6, column=2).value = res_samples["pw"]

# write the unsweeped vsl or vbl
if len(vbl_sweep) == 1:
    ws.cell(row=7, column=1).value = "vbl"
    ws.cell(row=7, column=2).value = vbl_sweep[0]
    v_sweep_name = "vwl/vsl"
    vbl_or_vsl_sweep = vsl_sweep
    get_bias = lambda vsl, vwl: (vbl_sweep[0], vsl, vwl)
else: # assume len(vsl_sweep) == 1:
    ws.cell(row=7, column=1).value = "vsl"
    ws.cell(row=7, column=2).value = vsl_sweep[0]
    v_sweep_name = "vwl/vbl"
    vbl_or_vsl_sweep = vbl_sweep
    get_bias = lambda vbl, vwl: (vbl, vsl_sweep[0], vwl)

# track current row

# MEAN TABLE
c0 = 2 # track initial table top-left col position
r0 = 10 # track table top-left row position
ws.cell(row=r0-1, column=1).value = "median"
ws.cell(row=r0, column=1).value = v_sweep_name
# header with vbl/vsl sweep value
for i, v in enumerate(vbl_or_vsl_sweep):
    ws.cell(row=r0, column=c0+i).value = v
# vwl column on left
r0 += 1
for j, vwl in enumerate(vwl_sweep):
    ws.cell(row=r0+j, column=c0-1).value = vwl
# write res mean table (separate loop for clarity)
for j, vwl in enumerate(vwl_sweep):
    for i, v in enumerate(vbl_or_vsl_sweep):
        bias = get_bias(v, vwl)
        if bias in res_samples["res"]:
            ws.cell(row=r0+j, column=c0+i).value = res_samples["res"][bias]["median"]

# STDDEV TABLE
c0 = 2 # track initial table top-left col position
r0 = r0 + 2 + len(vwl_sweep) # track table top-left row position
ws.cell(row=r0-1, column=1).value = "stddev"
ws.cell(row=r0, column=1).value = v_sweep_name
# header with vbl/vsl sweep value
for i, v in enumerate(vbl_or_vsl_sweep):
    ws.cell(row=r0, column=c0+i).value = v
# vwl column on left
r0 += 1
for j, vwl in enumerate(vwl_sweep):
    ws.cell(row=r0+j, column=c0-1).value = vwl
# write table (separate loop for clarity)
for j, vwl in enumerate(vwl_sweep):
    for i, v in enumerate(vbl_or_vsl_sweep):
        bias = get_bias(v, vwl)
        if bias in res_samples["res"]:
            ws.cell(row=r0+j, column=c0+i).value = res_samples["res"][bias]["std"]


wb.save(os.path.join(path_data_folder, "res_samples_statistics.xlsx"))
